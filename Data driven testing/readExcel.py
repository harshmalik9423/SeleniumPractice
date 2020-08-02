import openpyxl

path = r"C:\Users\harsh\Desktop\test.xlsx"      # r for reading as raw string
workbook = openpyxl.load_workbook(path)

# sheet = workbook.get_sheet_by_name("Sheet1")      # Extracts by name
sheet = workbook.active         # extracts active sheet

print("Rows = ", sheet.max_row)
print("Columns = ", sheet.max_column)

for r in range(1, sheet.max_row+1):
    for c in range(1, sheet.max_column+1):
        print(sheet.cell(row=r, column=c).value, end="    ")
    print()